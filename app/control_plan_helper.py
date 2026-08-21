"""Control-plan extraction, normalization, and quality checks."""
from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
import hashlib
import json
import os
import re

from app.ai_helper import OLLAMA_MODEL, _call_ollama, _parse_json


PARSER_VERSION = "cp-parser-1.2"
SUPPORTED_SPREADSHEETS = {"xlsx", "xlsm", "xls"}
SUPPORTED_PDFS = {"pdf"}


def sha256_file(file_path):
    digest = hashlib.sha256()
    with open(file_path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\r\n?", "\n", str(value)).strip()


def _norm(value):
    return re.sub(r"[\s_\-—–:：/\\()\[\]（）]+", "", _text(value).lower())


def _join_unique(values):
    output = []
    seen = set()
    for value in values:
        text = _text(value)
        if text and text not in seen:
            output.append(text)
            seen.add(text)
    return " ".join(output)


def _read_xlsx(file_path):
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    workbook = load_workbook(file_path, data_only=True, read_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        max_row = min(sheet.max_row or 1, 2000)
        max_col = min(sheet.max_column or 1, 80)
        matrix = [
            [_text(sheet.cell(row=row, column=col).value) for col in range(1, max_col + 1)]
            for row in range(1, max_row + 1)
        ]
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_merged_col, max_merged_row = range_boundaries(str(merged_range))
            if min_row > max_row or min_col > max_col:
                continue
            anchor = matrix[min_row - 1][min_col - 1]
            for row in range(min_row, min(max_merged_row, max_row) + 1):
                for col in range(min_col, min(max_merged_col, max_col) + 1):
                    matrix[row - 1][col - 1] = anchor
        sheets.append({"name": sheet.title, "matrix": matrix})
    workbook.close()
    return sheets


def _read_xls(file_path):
    import xlrd

    workbook = xlrd.open_workbook(file_path, formatting_info=False)
    sheets = []
    for sheet in workbook.sheets():
        max_row = min(sheet.nrows, 2000)
        max_col = min(sheet.ncols, 80)
        matrix = [
            [_text(sheet.cell_value(row, col)) for col in range(max_col)]
            for row in range(max_row)
        ]
        for min_row, max_row_exclusive, min_col, max_col_exclusive in sheet.merged_cells:
            if min_row >= max_row or min_col >= max_col:
                continue
            anchor = matrix[min_row][min_col]
            for row in range(min_row, min(max_row_exclusive, max_row)):
                for col in range(min_col, min(max_col_exclusive, max_col)):
                    matrix[row][col] = anchor
        sheets.append({"name": sheet.name, "matrix": matrix})
    return sheets


def _read_pdf(file_path):
    """Merge ruled control-plan tables from every PDF page into one matrix."""
    import pdfplumber

    matrix = []
    page_text = []
    table_page_count = 0
    page_count = 0
    expected_columns = None

    def table_candidates(tables):
        candidates = []
        for table in tables or []:
            if not table:
                continue
            column_count = max((len(row or []) for row in table), default=0)
            populated = sum(
                bool(_text(cell))
                for row in table
                for cell in (row or [])
            )
            if column_count >= 4 and populated:
                candidates.append((populated, column_count, table))
        return candidates

    with pdfplumber.open(file_path) as document:
        page_count = len(document.pages)
        for page_number, page in enumerate(document.pages, start=1):
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            if text:
                page_text.append(text)

            default_candidates = table_candidates(page.extract_tables())
            selected = max(default_candidates, key=lambda item: item[:2]) if default_candidates else None
            if expected_columns is None and selected:
                expected_columns = selected[1]

            verticals = sorted({
                round(line["x0"], 1)
                for line in page.lines
                if abs(line["x0"] - line["x1"]) < 1
                and abs(line["bottom"] - line["top"]) > page.height * 0.08
            })
            horizontals = sorted({
                round(line["top"], 1)
                for line in page.lines
                if abs(line["top"] - line["bottom"]) < 1
                and abs(line["x1"] - line["x0"]) > page.width * 0.05
            })

            # Keep pdfplumber's merged-cell interpretation when the page has
            # the expected shape. Only reconstruct the ruled grid on pages
            # where a large vertical merge collapses columns (common in long
            # supplier control plans that continue across pages).
            needs_grid_recovery = (
                selected is None
                or (expected_columns and selected[1] != expected_columns)
            )
            if needs_grid_recovery and len(verticals) >= 8 and len(horizontals) >= 2:
                recovered = table_candidates(page.extract_tables({
                    "vertical_strategy": "explicit",
                    "explicit_vertical_lines": verticals,
                    "horizontal_strategy": "explicit",
                    "explicit_horizontal_lines": horizontals,
                    "intersection_tolerance": 5,
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                }))
                if recovered:
                    recovered_best = max(recovered, key=lambda item: item[:2])
                    if (
                        selected is None
                        or recovered_best[1] == expected_columns
                        or abs(recovered_best[1] - expected_columns)
                        < abs(selected[1] - expected_columns)
                    ):
                        selected = recovered_best

            if selected is None:
                continue

            # Supplier PDFs commonly contain one full-width process table per page.
            _, column_count, table = selected
            table_page_count += 1
            for row in table:
                # pdfplumber uses None for cells covered by a vertical merge and
                # an empty string for a genuinely blank cell. Preserve that
                # distinction so the parser can inherit merged control fields.
                normalized = [
                    None if cell is None else _text(cell)
                    for cell in (row or [])
                ]
                normalized.extend([None] * (column_count - len(normalized)))
                matrix.append(normalized)

    return [{
        "name": f"PDF pages 1-{page_count}",
        "matrix": matrix,
        "metadata_text": "\n".join(page_text[:3]),
        "page_count": page_count,
        "table_page_count": table_page_count,
    }]


HEADER_ALIASES = {
    "process_code": [
        "part/processnumber", "partprocessnumber", "零件/过程编号", "零件过程编号",
    ],
    "process_name": [
        "processname/operationdescription", "processnameoperationdescription",
        "过程名称/操作描述", "工序名称", "processname",
    ],
    "machine": [
        "productionequipment", "生产设备", "machine/equipment", "设备/工装",
        "机器设备工装夹具", "machinedevicejigtoolformfg", "machinedevicejigtool",
        "machine device.jig. tool for mfg", "machine.device.jig.toolformfg",
    ],
    "char_code": ["serialnumber", "特性编号", "序号"],
    "product_char": [
        "characteristicproductitem", "产品项目", "productitem",
        "特性characteristics产品product", "特性产品", "产品product",
    ],
    "process_char": [
        "characteristicprocess", "过程特性", "processcharacteristic", "过程process",
        "特性characteristics过程process", "特性过程",
    ],
    "special_class": [
        "classificationofspecialcharacteristics", "特殊特性分类", "specialcharacteristic",
    ],
    "specification": [
        "product/processspecifications/tolerances", "productprocessspecificationstolerances",
        "产品/过程规范/公差", "规格/公差", "specification/tolerance",
    ],
    "measurement_method": [
        "evaluation/measurementtechnology", "evaluationmeasurementtechnology",
        "评价/测量技术", "测量方法", "检测方法",
    ],
    "sample_size": ["samplecapacity", "样本容量", "sample size", "samplesize", "capacity", "容量"],
    "frequency": ["samplefrequency", "检验频率", "频次", "frequency"],
    "inspector": [
        "sampleinspectors", "检验人员", "inspectors", "负责人", "actionholder",
    ],
    "control_method": ["controlmethod", "控制方法", "控制方式"],
    "reaction_plan": ["adversereactions", "reactionplan", "不良反应", "反应计划"],
    "step_no": ["no.", "序号", "编号"],
    "process_record": ["过程记录", "processrecord"],
    "key_parameters": ["关键参数", "keyparameters"],
}

DOCUMENT_HEADER_LABELS = {
    _norm(label)
    for label in (
        "production stage", "生产阶段",
        "control plan number", "控制计划编号",
        "part number/latest change level", "零件编号/最新更改等级",
        "part name/description", "零件名称", "零件代号",
        "organization/factory", "供方/工厂名称", "供方代码",
        "key contact", "主要联系人", "core team", "核心小组",
        "organization/factory approval/date", "供方/工厂批准/日期",
        "customer engineering approval/date", "顾客工程批准/日期",
        "customer quality approval/date", "顾客质量批准/日期",
        "other approval/date", "其它批准/日期",
        "date (compilation)", "编制日期",
        "part/process number", "零件/过程编号",
        "process name/operation description", "过程名称/操作描述",
        "machine, device, jig, tooling", "机器、装置夹具、工装",
        "characteristic", "特性", "product", "产品", "process", "过程",
        "classification of special characteristics", "特殊特性分类",
        "product/process specifications/tolerances", "产品/过程标准/公差",
        "evaluation/measurement technology", "评价测量技术",
        "sample", "样本", "capacity", "容量", "frequency", "频率",
        "control method", "控制方法", "reaction plan", "反应计划",
    )
}


def _column_mapping(headers):
    mapping = {}
    normalized = [_norm(header) for header in headers]
    for semantic, aliases in HEADER_ALIASES.items():
        best = None
        for index, header in enumerate(normalized):
            for alias in aliases:
                alias_norm = _norm(alias)
                if alias_norm and alias_norm in header:
                    candidate = (len(alias_norm), index)
                    if best is None or candidate > best:
                        best = candidate
        if best is not None:
            mapping[semantic] = best[1]
    return mapping


def _find_header(matrix):
    best = {"kind": "unknown", "score": 0, "start": 0, "height": 1, "mapping": {}}
    max_header_row = min(len(matrix), 45)
    max_col = max((len(row) for row in matrix[:max_header_row]), default=0)
    for start in range(max_header_row):
        for height in (1, 2, 3):
            if start + height > len(matrix):
                continue
            headers = [
                _join_unique(
                    matrix[row][col] if col < len(matrix[row]) else ""
                    for row in range(start, start + height)
                )
                for col in range(max_col)
            ]
            mapping = _column_mapping(headers)
            standard_keys = {
                "process_name", "specification", "measurement_method",
                "sample_size", "frequency", "control_method", "reaction_plan",
            }
            standard_score = len(standard_keys.intersection(mapping)) * 2
            if "process_name" in mapping and (
                "product_char" in mapping or "process_char" in mapping
            ):
                standard_score += 5
            if "process_code" in mapping:
                standard_score += 2
            generic_score = len(
                {"process_name", "process_record", "key_parameters"}.intersection(mapping)
            ) * 4
            if "process_name" in mapping and "step_no" in mapping:
                generic_score += 1
            if standard_score > best["score"]:
                best = {
                    "kind": "aiag",
                    "score": standard_score,
                    "start": start,
                    "height": height,
                    "mapping": mapping,
                }
            if generic_score > best["score"]:
                best = {
                    "kind": "process_record",
                    "score": generic_score,
                    "start": start,
                    "height": height,
                    "mapping": mapping,
                }
    return best


def _sheet_score(sheet):
    matrix = sheet["matrix"]
    header = _find_header(matrix)
    score = header["score"]
    name = _norm(sheet["name"])
    if "flowchart" in name or "流程图" in name:
        score -= 8
    populated = sum(1 for row in matrix[:80] if any(_text(cell) for cell in row))
    score += min(populated / 20, 4)
    return score, header


def _extract_metadata(matrix, extra_text=""):
    cells = []
    seen = set()
    for row in matrix[:14]:
        for cell in row[:24]:
            value = _text(cell)
            if value and value not in seen:
                cells.append(value)
                seen.add(value)
    # The PDF text layer preserves cover metadata better than a ruled-table
    # extraction, where values may be split across adjacent cells.
    text = extra_text.strip()
    if cells:
        cell_text = "\n".join(cells)
        text = f"{text}\n{cell_text}" if text else cell_text

    def match(patterns):
        for pattern in patterns:
            result = re.search(pattern, text, flags=re.IGNORECASE)
            if result:
                candidate = _text(result.group(1)).split("\n")[0].strip(" :：")
                if _norm(candidate) in {
                    "partno", "partname", "suppliernamecode", "cpno",
                    "preparedbydate", "approvalbydate", "reviseddate",
                }:
                    continue
                return candidate
        return ""

    return {
        "control_plan_number": match([
            r"Control\s*Plan\s*Number\s*[:：]\s*([A-Za-z0-9._/-]+)",
            r"控制计划编号\s*[:：]\s*([A-Za-z0-9._/-]+)",
            r"控制计划号码\s*[:：]\s*([A-Za-z0-9._/-]+)",
        ]),
        "part_number": match([
            r"零件号码\s*[:：]\s*([A-Za-z0-9._/-]+)",
            r"Part\s*Number(?:/Latest\s*Change\s*Level)?\s*[:：]\s*([^\n]+)",
            r"零件号(?:/最新更改水平)?\s*[:：]\s*([^\n]+)",
        ]),
        "part_name": match([
            r"零件名称\s*[:：]\s*(.+?)(?=\s*车型年度|\s*Model\s*Year|\n|$)",
            r"Part\s*Name/Description\s*[:：]\s*([^\n]+)",
            r"零件名称/描述\s*[:：]\s*([^\n]+)",
        ]),
        "organization": match([
            r"供应商名称/代码\s*[:：]\s*(.*?)(?=\s*控制计划号码|\s*CP\s*No|\n|$)",
            r"Organization/Factory\s*[:：]\s*([^\n]+)",
            r"组织/工厂\s*[:：]\s*([^\n]+)",
        ]),
        "compiled_date": match([
            r"编制/日期\s*[:：]\s*(.+?)(?=\s*供应商批准|\s*Approval\s*by|\s*顾客批准|\s*修订日期|\n|$)",
            r"Date\s*\(compilation\)\s*[:：]\s*([^\n]+)",
            r"日期\s*[（(]编制[）)]\s*[:：]\s*([^\n]+)",
        ]),
        "revised_date": match([
            r"Date\s*\(Revised\)\s*[:：]\s*([^\n]+)",
            r"日期\s*[（(]修订[）)]\s*[:：]\s*([^\n]+)",
            r"修订日期\s*[:：]\s*([^\n]+)",
        ]),
    }


def _row_value(row, mapping, key):
    index = mapping.get(key)
    if index is None or index >= len(row):
        return ""
    return _text(row[index])


def _is_header_like(value):
    normalized = _norm(value)
    return normalized in {
        "productitem", "产品项目", "process", "过程", "characteristic",
        "特性", "processnameoperationdescription", "工序名称",
    }


def _is_document_header_row(row, mapping):
    values = {
        key: _row_value(row, mapping, key)
        for key in (
            "process_code", "process_name", "machine", "char_code",
            "product_char", "process_char", "special_class", "specification",
            "measurement_method", "sample_size", "frequency", "inspector",
            "control_method", "reaction_plan",
        )
    }
    normalized = {key: _norm(value) for key, value in values.items() if value}
    hits = {
        value for value in normalized.values()
        if value in DOCUMENT_HEADER_LABELS
    }
    process_code = normalized.get("process_code", "")
    process_name = normalized.get("process_name", "")

    if process_code in DOCUMENT_HEADER_LABELS and (
        not process_name
        or process_name == process_code
        or process_name in DOCUMENT_HEADER_LABELS
    ):
        return True
    return len(hits) >= 2


def _parse_aiag(matrix, header, sheet_name, confidence=0.95, pdf_mode=False):
    mapping = header["mapping"]
    steps_by_key = OrderedDict()
    signatures_by_key = {}
    active_key = None
    inherited = {}
    start = header["start"] + header["height"]
    for source_index, row in enumerate(matrix[start:], start=start + 1):
        if _is_document_header_row(row, mapping):
            inherited = {}
            active_key = None
            continue
        process_code = _row_value(row, mapping, "process_code")
        process_name = _row_value(row, mapping, "process_name")
        machine = _row_value(row, mapping, "machine")
        product_char = _row_value(row, mapping, "product_char")
        process_char = _row_value(row, mapping, "process_char")
        char_code = _row_value(row, mapping, "char_code")
        specification = _row_value(row, mapping, "specification")
        measurement = _row_value(row, mapping, "measurement_method")
        control_method = _row_value(row, mapping, "control_method")
        reaction = _row_value(row, mapping, "reaction_plan")

        if _is_header_like(product_char) or _is_header_like(process_char):
            continue
        if not any((
            process_code, process_name, product_char, process_char, specification,
            measurement, control_method, reaction,
        )):
            continue

        section_only = process_code and not any((
            process_name, machine, product_char, process_char, specification,
            measurement, control_method, reaction,
        ))
        compact_code = re.sub(r"\s+", "", process_code)
        if section_only and not re.fullmatch(r"[A-Za-z0-9._/-]{1,24}", compact_code):
            inherited = {}
            active_key = None
            continue

        # Explicit PDF grid lines can split text from a vertically merged
        # process cell across two characteristic rows. Only the row carrying
        # the process code should declare that process.
        if (
            pdf_mode and active_key is not None and not process_code
            and char_code not in {"", "1"}
            and any((product_char, process_char, specification))
        ):
            process_name = ""
            machine = ""

        declares_step = any((process_code, process_name, machine))

        def inherit_merged(key, value):
            index = mapping.get(key)
            if value:
                inherited[key] = value
                return value
            if index is not None and index < len(row) and (
                row[index] is None or (pdf_mode and not _text(row[index]))
            ):
                return inherited.get(key, "")
            return value

        machine = inherit_merged("machine", machine)
        measurement = inherit_merged("measurement_method", measurement)
        control_method = inherit_merged("control_method", control_method)
        reaction = inherit_merged("reaction_plan", reaction)

        if declares_step:
            key = (process_code, process_name, machine)
            active_key = key
        elif active_key is not None:
            key = active_key
        else:
            key = ("", "", "")
        if key not in steps_by_key:
            steps_by_key[key] = {
                "seq": (len(steps_by_key) + 1) * 10,
                "process_code": key[0],
                "process_name": key[1] or key[0] or "Unspecified process",
                "machine": key[2],
                "is_key_process": False,
                "notes": "",
                "source_sheet": sheet_name,
                "source_row": source_index,
                "characteristics": [],
            }
            signatures_by_key[key] = set()

        char_name = product_char or process_char
        char_type = "product" if product_char else "process"
        if not char_name and specification:
            char_name = "Control characteristic"
            char_type = "process"
        if not char_name:
            continue

        special_class = inherit_merged(
            "special_class", _row_value(row, mapping, "special_class")
        )
        sample_size = inherit_merged(
            "sample_size", _row_value(row, mapping, "sample_size")
        )
        frequency = inherit_merged(
            "frequency", _row_value(row, mapping, "frequency")
        )
        inspector = inherit_merged(
            "inspector", _row_value(row, mapping, "inspector")
        )
        characteristic = {
            "char_code": char_code,
            "char_name": char_name,
            "char_type": char_type,
            "special_class": special_class,
            "spec_value": specification,
            "spec_unit": "",
            "tolerance": "",
            "measurement_method": measurement,
            "sample_size": sample_size,
            "frequency": frequency,
            "inspector": inspector,
            "control_method": control_method,
            "reaction_plan": reaction,
            "is_key_char": bool(special_class.strip()),
            "source_sheet": sheet_name,
            "source_row": source_index,
            "confidence": confidence,
        }
        signature = tuple(
            characteristic.get(field, "")
            for field in (
                "char_code", "char_name", "char_type", "special_class",
                "spec_value", "measurement_method", "sample_size", "frequency",
                "inspector", "control_method", "reaction_plan",
            )
        )
        if signature in signatures_by_key[key]:
            continue
        signatures_by_key[key].add(signature)
        steps_by_key[key]["characteristics"].append(characteristic)

    return list(steps_by_key.values())


def _split_parameters(value):
    lines = []
    for raw_line in re.split(r"[\n;；]+", _text(value)):
        line = raw_line.strip(" •·-\t")
        if line:
            lines.append(line)
    output = []
    for line in lines:
        parts = re.split(r"\s*[:：]\s*", line, maxsplit=1)
        if len(parts) == 2 and parts[0] and parts[1]:
            output.append((parts[0], parts[1]))
        else:
            output.append(("Key parameter", line))
    return output


def _parse_process_record(matrix, header, sheet_name):
    mapping = header["mapping"]
    steps = []
    start = header["start"] + header["height"]
    for source_index, row in enumerate(matrix[start:], start=start + 1):
        process_name = _row_value(row, mapping, "process_name")
        process_record = _row_value(row, mapping, "process_record")
        key_parameters = _row_value(row, mapping, "key_parameters")
        if not any((process_name, process_record, key_parameters)):
            continue
        if _is_header_like(process_name):
            continue
        is_key = "★" in process_name or "关键" in process_name
        process_name = process_name.replace("★", "").strip()
        step = {
            "seq": (len(steps) + 1) * 10,
            "process_code": _row_value(row, mapping, "step_no"),
            "process_name": process_name or "Unspecified process",
            "machine": "",
            "is_key_process": is_key,
            "notes": process_record,
            "source_sheet": sheet_name,
            "source_row": source_index,
            "characteristics": [],
        }
        for char_index, (name, specification) in enumerate(_split_parameters(key_parameters), start=1):
            step["characteristics"].append({
                "char_code": str(char_index),
                "char_name": name,
                "char_type": "process",
                "special_class": "Key process" if is_key else "",
                "spec_value": specification,
                "spec_unit": "",
                "tolerance": "",
                "measurement_method": "",
                "sample_size": "",
                "frequency": "",
                "inspector": "",
                "control_method": process_record,
                "reaction_plan": "",
                "is_key_char": is_key,
                "source_sheet": sheet_name,
                "source_row": source_index,
                "confidence": 0.78,
            })
        steps.append(step)
    return steps


def _ai_map_matrix(matrix, logger=None):
    rows = []
    for row_number, row in enumerate(matrix[:40], start=1):
        cells = [_text(cell)[:160] for cell in row[:20]]
        if any(cells):
            rows.append({"row": row_number, "cells": cells})
    prompt = f"""You map supplier control-plan spreadsheets into a fixed schema.
Return JSON only:
{{
  "header_row": 1,
  "header_rows": 1,
  "kind": "aiag" or "process_record",
  "columns": {{
    "process_code": 1, "process_name": 2, "machine": 3, "char_code": 4,
    "product_char": 5, "process_char": 6, "special_class": 7,
    "specification": 8, "measurement_method": 9, "sample_size": 10,
    "frequency": 11, "inspector": 12, "control_method": 13,
    "reaction_plan": 14, "step_no": 1, "process_record": 3,
    "key_parameters": 4
  }}
}}
Column and row numbers are 1-based. Omit fields that are absent. Do not invent data.
The header row must be the actual process-table column header directly above the
first process record. Do not use title, part information, supplier information,
approval, contact, revision, or other document metadata rows as the header.

Spreadsheet rows:
{json.dumps(rows, ensure_ascii=False)}
"""
    raw = _call_ollama(prompt, timeout=90, num_predict=600, logger=logger)
    parsed = _parse_json(raw)
    if not isinstance(parsed, dict) or not isinstance(parsed.get("columns"), dict):
        return None
    mapping = {}
    for key, value in parsed["columns"].items():
        if key in HEADER_ALIASES:
            try:
                index = int(value) - 1
            except (TypeError, ValueError):
                continue
            if index >= 0:
                mapping[key] = index
    if "process_name" not in mapping:
        return None
    try:
        start = max(0, int(parsed.get("header_row", 1)) - 1)
        height = max(1, min(3, int(parsed.get("header_rows", 1))))
    except (TypeError, ValueError):
        start, height = 0, 1
    kind = parsed.get("kind")
    if kind not in {"aiag", "process_record"}:
        kind = "process_record" if "key_parameters" in mapping else "aiag"
    return {"kind": kind, "start": start, "height": height, "mapping": mapping}


def _align_ai_header(ai_header, detected_header):
    """Keep AI column mapping but prefer a confidently detected table boundary."""
    if not ai_header or not detected_header:
        return ai_header
    if detected_header.get("score", 0) < 9:
        return ai_header
    aligned = dict(ai_header)
    aligned["start"] = detected_header["start"]
    aligned["height"] = detected_header["height"]
    return aligned


def assess_quality(data):
    steps = data.get("steps") or []
    issues = []

    def add(code, severity, count, message):
        if count:
            issues.append({
                "code": code,
                "severity": severity,
                "count": count,
                "message": message,
            })

    characteristics = [
        characteristic
        for step in steps
        for characteristic in step.get("characteristics", [])
    ]
    add(
        "no_steps", "critical", int(not steps),
        "未识别到工序，请检查原始表格或使用 AI 重新识别。",
    )
    add(
        "empty_steps", "warning",
        sum(not step.get("characteristics") for step in steps),
        "部分工序没有控制特性。",
    )
    add(
        "missing_spec", "critical",
        sum(not item.get("spec_value") for item in characteristics),
        "部分控制特性缺少规格或公差。",
    )
    add(
        "missing_measurement", "warning",
        sum(not item.get("measurement_method") for item in characteristics),
        "部分控制特性缺少测量方法。",
    )
    add(
        "missing_sampling", "warning",
        sum(not item.get("sample_size") or not item.get("frequency") for item in characteristics),
        "部分控制特性缺少抽样数量或频次。",
    )
    add(
        "missing_control", "warning",
        sum(not item.get("control_method") for item in characteristics),
        "部分控制特性缺少控制方法或记录依据。",
    )
    add(
        "missing_reaction", "warning",
        sum(not item.get("reaction_plan") for item in characteristics),
        "部分控制特性缺少超差反应计划。",
    )
    vague_pattern = re.compile(r"定期|适时|必要时|regularly|periodically|as needed", re.I)
    add(
        "vague_frequency", "warning",
        sum(bool(vague_pattern.search(item.get("frequency") or "")) for item in characteristics),
        "部分检验频次表述不够明确。",
    )
    add(
        "incomplete_key_characteristic", "critical",
        sum(
            item.get("is_key_char") and (
                not item.get("spec_value")
                or not item.get("measurement_method")
                or not item.get("reaction_plan")
            )
            for item in characteristics
        ),
        "关键特性的规格、测量方法或反应计划不完整。",
    )

    counts = {item["code"]: item["count"] for item in issues}
    score = 100
    score -= 30 if counts.get("no_steps") else 0
    score -= min(15, counts.get("empty_steps", 0) * 3)
    score -= min(25, counts.get("missing_spec", 0) * 2)
    score -= min(12, counts.get("missing_measurement", 0))
    score -= min(12, counts.get("missing_sampling", 0))
    score -= min(10, counts.get("missing_control", 0))
    score -= min(12, counts.get("missing_reaction", 0))
    score -= min(8, counts.get("incomplete_key_characteristic", 0) * 2)
    score -= min(5, counts.get("vague_frequency", 0))
    return max(0, score), issues


def extract_control_plan(file_path, force_ai=False, logger=None):
    extension = os.path.splitext(file_path)[1].lower().lstrip(".")
    base = {
        "parser_version": PARSER_VERSION,
        "source_sheet": "",
        "source_template": "unsupported",
        "confidence": 0.0,
        "ai_model": None,
        "metadata": {},
        "steps": [],
    }
    if extension not in SUPPORTED_SPREADSHEETS | SUPPORTED_PDFS:
        score, issues = assess_quality(base)
        issues.insert(0, {
            "code": "unsupported_format",
            "severity": "warning",
            "count": 1,
            "message": "原文件已保存；当前结构化提取支持 PDF、XLS 和 XLSX 控制计划。",
        })
        base.update({"quality_score": score, "quality_issues": issues})
        return base

    is_pdf = extension in SUPPORTED_PDFS
    if is_pdf:
        sheets = _read_pdf(file_path)
    elif extension == "xls":
        sheets = _read_xls(file_path)
    else:
        sheets = _read_xlsx(file_path)
    if not sheets:
        raise ValueError("附件中没有可读取的控制计划内容")

    if is_pdf and not sheets[0].get("matrix"):
        base.update({
            "source_sheet": sheets[0]["name"],
            "source_template": "pdf_no_table",
            "metadata": _extract_metadata([], sheets[0].get("metadata_text", "")),
        })
        score, issues = assess_quality(base)
        issues.insert(0, {
            "code": "pdf_no_table",
            "severity": "critical",
            "count": 1,
            "message": "PDF 中未检测到可解析表格；如为扫描件，请先进行 OCR 或上传原始 Excel。",
        })
        base.update({"quality_score": score, "quality_issues": issues})
        return base

    ranked = []
    for sheet in sheets:
        score, header = _sheet_score(sheet)
        ranked.append((score, sheet, header))
    _, selected_sheet, header = max(ranked, key=lambda item: item[0])
    matrix = selected_sheet["matrix"]
    ai_used = False

    if header["kind"] == "aiag" and header["score"] >= 10:
        steps = _parse_aiag(
            matrix, header, selected_sheet["name"],
            confidence=0.90 if is_pdf else 0.95,
            pdf_mode=is_pdf,
        )
        source_template = "pdf_aiag" if is_pdf else "aiag"
        confidence = 0.90 if is_pdf else 0.95
    elif header["kind"] == "process_record" and header["score"] >= 9:
        steps = _parse_process_record(matrix, header, selected_sheet["name"])
        source_template = "pdf_process_record" if is_pdf else "process_record"
        confidence = 0.74 if is_pdf else 0.78
    else:
        steps = []
        source_template = "unknown"
        confidence = 0.35

    # A clear ruled PDF table is more reliable than asking a text-only model to
    # guess its columns again. AI mapping remains the fallback for weak tables.
    if (force_ai and not is_pdf) or not steps:
        ai_header = _ai_map_matrix(matrix, logger=logger)
        if ai_header:
            ai_header = _align_ai_header(ai_header, header)
            if ai_header["kind"] == "process_record":
                ai_steps = _parse_process_record(matrix, ai_header, selected_sheet["name"])
                ai_template = "ai_process_record"
            else:
                ai_steps = _parse_aiag(
                    matrix, ai_header, selected_sheet["name"], confidence=0.82,
                    pdf_mode=is_pdf,
                )
                ai_template = "ai_mapped"
            current_count = sum(len(step["characteristics"]) for step in steps)
            ai_count = sum(len(step["characteristics"]) for step in ai_steps)
            ai_is_complete_enough = (
                not current_count or ai_count >= max(1, int(current_count * 0.6))
            )
            if ai_steps and (
                ai_count > current_count or (force_ai and ai_is_complete_enough)
            ):
                steps = ai_steps
                source_template = ai_template
                confidence = 0.82
                ai_used = True

    data = {
        "parser_version": PARSER_VERSION,
        "source_sheet": selected_sheet["name"],
        "source_template": source_template,
        "confidence": confidence,
        "ai_model": OLLAMA_MODEL if ai_used else None,
        "metadata": _extract_metadata(
            matrix, selected_sheet.get("metadata_text", "")
        ),
        "steps": steps,
    }
    quality_score, quality_issues = assess_quality(data)
    data["quality_score"] = quality_score
    data["quality_issues"] = quality_issues
    return data
