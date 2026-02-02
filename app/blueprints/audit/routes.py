"""
Audit Findings Routes
审核发现路由 - 支持 Excel 自动提取和 PDF 手动录入
"""

from flask import render_template, request, redirect, url_for, flash, send_file, abort, current_app
from werkzeug.utils import secure_filename
from sqlalchemy import or_
from datetime import datetime, date
import os
import uuid
import openpyxl
import pandas as pd

from . import audit_bp  # ← 从当前包导入 blueprint
from ...extensions import db  # ← 注意这里是三个点（上两级）
from ...models import AuditReport, AuditFinding, FindingProgress, FindingAttachment, Supplier

# 允许的文件扩展名 - 新增 PDF 支持
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm', 'pdf'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def generate_audit_no():
    """生成审核编号：AUD-YYYY-XXX"""
    current_year = datetime.now().year

    last_audit = AuditReport.query.filter(
        AuditReport.audit_no.like(f"AUD-{current_year}-%")
    ).order_by(AuditReport.audit_no.desc()).first()

    if last_audit:
        try:
            last_num = int(last_audit.audit_no.split("-")[-1])
            new_num = last_num + 1
        except (ValueError, IndexError):
            new_num = 1
    else:
        new_num = 1

    return f"AUD-{current_year}-{new_num:03d}"


def extract_anfia_findings(file_path, report):
    """
    从 ANFIA Excel 文件中提取问题点
    Extract findings from ANFIA Excel Action Plan sheet
    """
    findings = []

    try:
        # 读取 Excel 文件
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 查找 Action Plan sheet
        action_plan_sheet = None
        possible_names = ['Action Plan', 'ACTION PLAN', 'Action plan', 'action plan',
                          '行动计划', 'ActionPlan']

        for sheet_name in wb.sheetnames:
            if any(name in sheet_name for name in possible_names):
                action_plan_sheet = wb[sheet_name]
                break

        if not action_plan_sheet:
            flash("Warning: No 'Action Plan' sheet found in Excel file.", "warning")
            return findings

        # 从第 23 行开始读取数据（跳过标题和表头）
        for row_idx, row in enumerate(action_plan_sheet.iter_rows(min_row=23, values_only=True), start=23):
            # 跳过完全空的行
            if not any(row):
                continue

            # 提取条款号（A列）
            clause_no = str(row[0]).strip() if row[0] else ""
            if not clause_no or clause_no.lower() in ['none', 'n/a', '', 'null']:
                continue

            # 提取问题描述（B列）
            finding_text = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not finding_text or finding_text.lower() in ['none', 'n/a', '', 'null']:
                continue

            # 提取严重程度（C列）
            severity_raw = str(row[2]).strip().lower() if len(row) > 2 and row[2] else "minor"

            # 映射严重程度
            if 'major' in severity_raw or 'iii' in severity_raw or severity_raw == '3':
                severity = 'major'
            elif 'ii' in severity_raw or severity_raw == '2':
                severity = 'minor'
            elif 'i' in severity_raw or severity_raw == '1':
                severity = 'observation'
            else:
                severity = 'minor'

            # 提取纠正措施（D列）
            corrective_action = str(row[3]).strip() if len(row) > 3 and row[3] else None
            if corrective_action and corrective_action.lower() in ['none', 'n/a', '', 'null']:
                corrective_action = None

            # 提取负责人（E列）
            responsible_person = str(row[4]).strip() if len(row) > 4 and row[4] else None
            if responsible_person and responsible_person.lower() in ['none', 'n/a', '', 'null']:
                responsible_person = None

            # 提取目标日期（F列）
            target_date = None
            if len(row) > 5 and row[5]:
                try:
                    if isinstance(row[5], datetime):
                        target_date = row[5].date()
                    elif isinstance(row[5], date):
                        target_date = row[5]
                    elif isinstance(row[5], str):
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                            try:
                                target_date = datetime.strptime(str(row[5]).strip(), fmt).date()
                                break
                            except:
                                continue
                except Exception as e:
                    print(f"Date parsing error for row {row_idx}: {e}")
                    pass

            # 创建 Finding 对象
            finding = AuditFinding(
                report_id=report.id,
                clause_no=clause_no,
                requirement=f"Requirement {clause_no}",
                finding=finding_text,
                severity=severity,
                corrective_action=corrective_action,
                responsible_person=responsible_person,
                target_date=target_date,
                status='open'
            )

            findings.append(finding)
            print(f"✓ Extracted finding: {clause_no} - {finding_text[:50]}...")

        wb.close()
        print(f"Total findings extracted: {len(findings)}")

    except Exception as e:
        flash(f"Error extracting findings from Excel: {str(e)}", "error")
        print(f"Excel extraction error: {e}")
        import traceback
        traceback.print_exc()

    return findings


@audit_bp.route('/', methods=['GET'])
def index():
    """审核列表页"""
    q = (request.args.get('q') or '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20

    query = AuditReport.query

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                AuditReport.audit_no.ilike(like),
                AuditReport.supplier_name.ilike(like),
                AuditReport.auditor.ilike(like),
                AuditReport.audit_type.ilike(like)
            )
        )

    query = query.order_by(AuditReport.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    reports = pagination.items

    # 统计数据
    stats = {
        'total_reports': AuditReport.query.count(),
        'open_findings': AuditFinding.query.filter_by(status='open').count(),
        'in_progress': AuditFinding.query.filter_by(status='in_progress').count(),
        'overdue': AuditFinding.query.filter(
            AuditFinding.target_date < date.today(),
            AuditFinding.status.in_(['open', 'in_progress'])
        ).count()
    }

    return render_template(
        'audit/index.html',
        reports=reports,
        pagination=pagination,
        stats=stats,
        q=q
    )


@audit_bp.route('/upload', methods=['GET', 'POST'])
def upload_report():
    """上传审核报告 - 支持 Excel 和 PDF"""
    if request.method == 'POST':
        # 验证文件
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        # 检查文件类型
        original_filename = secure_filename(file.filename)
        file_ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''

        if file_ext not in ALLOWED_EXTENSIONS:
            flash(f'Invalid file type. Please upload Excel (.xlsx, .xls, .xlsm) or PDF (.pdf) file', 'error')
            return redirect(request.url)

        # 获取表单数据
        audit_type = request.form.get('audit_type', 'ANFIA')
        audit_date_str = request.form.get('audit_date')
        supplier_name = request.form.get('supplier_name')
        auditor = request.form.get('auditor')
        notes = request.form.get('notes', '').strip() or None

        # 验证必填字段
        if not all([audit_date_str, supplier_name, auditor]):
            flash('Please fill in all required fields', 'error')
            return redirect(request.url)

        try:
            audit_date = datetime.strptime(audit_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format', 'error')
            return redirect(request.url)

        # 生成审核编号
        audit_no = generate_audit_no()

        # 保存文件
        stored_filename = f"{uuid.uuid4().hex}.{file_ext}"

        upload_dir = os.path.join(current_app.config['UPLOAD_DIR'], 'audit_reports')
        os.makedirs(upload_dir, exist_ok=True)

        file_path = os.path.join(upload_dir, stored_filename)
        file.save(file_path)

        # 创建审核记录
        report = AuditReport(
            audit_no=audit_no,
            audit_type=audit_type,
            supplier_name=supplier_name,
            audit_date=audit_date,
            auditor=auditor,
            original_filename=original_filename,
            stored_filename=stored_filename,
            file_path=os.path.join('audit_reports', stored_filename),
            notes=notes,
            status='open'
        )

        db.session.add(report)
        db.session.flush()  # 获取 report.id

        # 根据文件类型决定是否自动提取
        findings = []
        if file_ext in ['xlsx', 'xls', 'xlsm'] and audit_type == 'ANFIA':
            # Excel 文件 - 自动提取
            findings = extract_anfia_findings(file_path, report)

            if findings:
                for finding in findings:
                    db.session.add(finding)

                report.total_findings = len(findings)
                report.open_findings = len(findings)
                report.closed_findings = 0

                flash(f'✅ Report uploaded successfully! Extracted {len(findings)} findings.', 'success')
            else:
                flash('⚠️ Report uploaded but no findings were extracted. You can add them manually.', 'warning')
        else:
            # PDF 文件或其他 - 手动添加
            flash(f'✅ Report uploaded successfully: {audit_no}. Please add findings manually.', 'success')

        db.session.commit()

        return redirect(url_for('audit.report_detail', report_id=report.id))

    # GET - 显示上传表单
    suppliers = Supplier.query.order_by(Supplier.code).all()
    return render_template('audit/upload.html', suppliers=suppliers)


@audit_bp.route('/report/<int:report_id>', methods=['GET'])
def report_detail(report_id):
    """审核报告详情"""
    report = AuditReport.query.get_or_404(report_id)

    # 获取筛选参数
    status_filter = request.args.get('status')

    query = report.findings

    if status_filter:
        query = query.filter_by(status=status_filter)

    findings = query.order_by(AuditFinding.clause_no).all()

    return render_template(
        'audit/detail.html',
        report=report,
        findings=findings
    )


@audit_bp.route('/report/<int:report_id>/add-finding', methods=['POST'])
def add_finding(report_id):
    """手动添加 Finding"""
    report = AuditReport.query.get_or_404(report_id)

    # 获取表单数据
    clause_no = (request.form.get('clause_no') or '').strip()
    clause_title = (request.form.get('clause_title') or '').strip() or None
    finding_text = (request.form.get('finding') or '').strip()
    evidence = (request.form.get('evidence') or '').strip() or None
    severity = (request.form.get('severity') or 'observation').strip()

    target_date_str = (request.form.get('target_date') or '').strip()
    responsible_person = (request.form.get('responsible_person') or '').strip() or None

    # 验证必填字段
    if not clause_no:
        flash('Clause No. is required', 'error')
        return redirect(url_for('audit.report_detail', report_id=report_id))

    if not finding_text:
        flash('Finding description is required', 'error')
        return redirect(url_for('audit.report_detail', report_id=report_id))

    # 解析目标日期
    target_date = None
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # 创建 Finding
    finding = AuditFinding(
        report_id=report.id,
        clause_no=clause_no,
        clause_title=clause_title,
        finding=finding_text,
        evidence=evidence,
        severity=severity,
        target_date=target_date,
        responsible_person=responsible_person,
        status='open'
    )

    db.session.add(finding)

    # 更新报告统计
    report.update_statistics()

    db.session.commit()

    flash(f'✅ Finding added successfully: {clause_no}', 'success')
    return redirect(url_for('audit.report_detail', report_id=report_id))


@audit_bp.route('/finding/<int:finding_id>/update', methods=['POST'])
def update_finding(finding_id):
    """更新问题点"""
    finding = AuditFinding.query.get_or_404(finding_id)

    old_status = finding.status

    # 更新字段
    finding.root_cause = request.form.get('root_cause', '').strip() or None
    finding.corrective_action = request.form.get('corrective_action', '').strip() or None
    finding.preventive_action = request.form.get('preventive_action', '').strip() or None
    finding.responsible_person = request.form.get('responsible_person', '').strip() or None
    finding.status = request.form.get('status', 'open')
    finding.verification_result = request.form.get('verification_result', '').strip() or None

    # 目标日期
    target_date_str = request.form.get('target_date', '').strip()
    if target_date_str:
        try:
            finding.target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # 如果状态改为 closed，记录完成日期
    if finding.status == 'closed' and old_status != 'closed':
        finding.actual_completion_date = date.today()
        finding.verification_date = date.today()

    # 记录进展更新
    comment = request.form.get('comment', '').strip()
    if comment or old_status != finding.status:
        progress = FindingProgress(
            finding_id=finding.id,
            update_type='status_change' if old_status != finding.status else 'supplier_update',
            old_status=old_status,
            new_status=finding.status,
            comment=comment,
            updated_by=finding.report.auditor
        )
        db.session.add(progress)

    db.session.commit()

    # 更新报告统计
    finding.report.update_statistics()
    db.session.commit()

    flash('✅ Finding updated successfully', 'success')
    return redirect(url_for('audit.report_detail', report_id=finding.report_id))


@audit_bp.route('/report/<int:report_id>/download', methods=['GET'])
def download_report(report_id):
    """下载原始审核报告"""
    report = AuditReport.query.get_or_404(report_id)

    file_path = os.path.join(current_app.config['UPLOAD_DIR'], report.file_path)

    if not os.path.exists(file_path):
        abort(404, 'File not found')

    return send_file(
        file_path,
        as_attachment=True,
        download_name=report.original_filename
    )


@audit_bp.route('/report/<int:report_id>/delete', methods=['POST'])
def delete_report(report_id):
    """删除审核报告"""
    report = AuditReport.query.get_or_404(report_id)

    # 删除文件
    file_path = os.path.join(current_app.config['UPLOAD_DIR'], report.file_path)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except OSError:
            pass

    db.session.delete(report)
    db.session.commit()

    flash('✅ Audit report deleted successfully', 'success')
    return redirect(url_for('audit.index'))